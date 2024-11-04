from openpyxl import load_workbook
from datetime import datetime

# Load the existing workbook and select the worksheet
file_path = '/Users/mihirdhankani/Downloads/LunchRecord.xlsx'
workbook = load_workbook(filename=file_path)
sheet = workbook.active  # Or specify sheet name with workbook['SheetName']

# Find the next available row in a specific column (e.g., column A)
column = input("Enter GR")
next_row = sheet.max_row + 1  # Assumes no gaps in column data

# Add today's date in the next available cell in the column
today_date = datetime.today().strftime('%Y-%m-%d')  # Format as YYYY-MM-DD
sheet[f'{column}{next_row}'] = today_date

# Save the workbook
workbook.save(filename=file_path)
print(
    f"Today's date ({today_date}) has been added to column {column} in row {next_row}.")
